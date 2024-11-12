import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import logging
from datetime import datetime
import os

class CSVtoExcelConverter:
    def __init__(self, csv_path: str = 'products.csv', excel_template_path: str = 'sample.xlsx'):
        self.csv_path = csv_path
        self.excel_template_path = excel_template_path
        self.logger = logging.getLogger(__name__)
        
        # 필드 매핑 정의 (엑셀 열 번호는 1부터 시작)
        self.field_mapping = {
            # 필수 필드
            'category_id': 1,             # A열: 메이크샵 카테고리
            'name': 4,                    # D열: 상품명
            'display_name': 5,            # E열: 모바일 상품명
            'option_mandatory': 6,        # F열: 옵션 필수 여부
            'option_mix': 7,              # G열: 옵션 조합 여부
            'option_name': 8,             # H열: 옵션명
            'option_values': 9,           # I열: 옵션값
            'option_prices': 10,          # J열: 옵션 가격
            'opt_use': 15,                # O열: 사용여부
            'opt_oneclick': 16,           # P열: 옵션 한번클릭 여부
            'option_type': 30,            # AD열: 옵션 출력 방식
            'selling_price': 31,          # AE열: 판매가격
            'stock': 32,                  # AF열: 재고수량
            'consumer_price': 33,         # AG열: 소비자가격
            'supply_price': 34,           # AH열: 상품 구매원가
            'point_percentage': 36,       # AI열: 적립금 비율
            'reserve': 37,                # AI열: 적립금
            'mobile_reserve': 38,         # AJ열: 모바일 적립금
            'point': 38,                  # AK열: 포인트
            'search_tags': 39,            # AM열: 키워드
            'adult_auth': 40,             # AM열: 판매가능여부
            'display_status': 41,         # AN열: 상품진열여부
            'main_image': 42,             # AO열: 확대이미지
            'detail_image': 43,           # AP열: 상세이미지
            'list_image': 44,             # AQ열: 리스트이미지
            'mobile_image': 46,           # AR열: 모바일 이미지
            'description': 45,            # AS열: 상품상세
            'mobile_description': 46,     # AT열: 모바일 상세설명
            'model_name': 57,             # BD열: 모델명
            'best_product_display': 62,   # BE열: 베스트상품 진열여부
            'vat_type': 64                # BG열: 부가세 설정
        }
        
        # 기본값 설정
        self.default_values = {
            'option_type': '분리형',
            'stock': 100,
            'display_status': '진열',
            'detail_image': 'AUTO',
            'list_image': 'AUTO',
            'adult_auth': 'Y',
            'point_percentage': '10%',
            'point': '',
            'best_product_display': 'N',
            'vat_type': 'Y',
            'option_mandatory': '선택',
            'option_mix': '미조합'
        }

    def read_csv(self) -> pd.DataFrame:
        """CSV 파일 읽기"""
        try:
            df = pd.read_csv(self.csv_path, encoding='utf-8-sig')
            return df
        except Exception as e:
            self.logger.error(f"CSV 파일 읽기 실패: {str(e)}")
            return pd.DataFrame()

    def prepare_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """데이터 전처리"""
        try:
            processed_df = pd.DataFrame()
            
            # 1. 기본 정보
            processed_df['category_id'] = df.iloc[:, 0]
            processed_df['name'] = df.iloc[:, 1]
            processed_df['display_name'] = df.iloc[:, 2]
            
            # 2. 옵션 관련 처리
            option_values = df.iloc[:, 10]  # option_values 열 (실제 옵션값이 있는 컬럼 위치로 수정)

            # 옵션 기본값 설정 (모두 빈 값으로 초기화)
            processed_df['option_mandatory'] = ''
            processed_df['option_mix'] = ''
            processed_df['option_name'] = ''
            processed_df['option_values'] = ''
            processed_df['option_prices'] = ''
            processed_df['opt_use'] = ''
            processed_df['opt_oneclick'] = ''
            processed_df['option_type'] = ''

            # 각 행의 옵션 처리
            for idx in df.index:
                opt_val = str(option_values[idx]).strip()
                # 단일상품이 아니고, 실제 옵션값이 있는 경우만 처리
                if opt_val and opt_val != '단일상품':
                    try:
                        # '|' 구분자를 ',' 구분자로 변경
                        options = [opt.strip() for opt in opt_val.split('|') if opt.strip()]
                        if options:  # 옵션값이 실제로 존재하는 경우만
                            processed_df.at[idx, 'option_mandatory'] = '필수'
                            processed_df.at[idx, 'option_mix'] = '미조합'
                            processed_df.at[idx, 'option_name'] = '기타'
                            processed_df.at[idx, 'option_values'] = ','.join(options)
                            processed_df.at[idx, 'option_prices'] = ','.join(['0'] * len(options))
                            processed_df.at[idx, 'opt_use'] = '사용'
                            processed_df.at[idx, 'opt_oneclick'] = '사용'
                            processed_df.at[idx, 'option_type'] = '분리형'
                    except Exception as e:
                        self.logger.error(f"옵션 처리 중 오류 (행 {idx}): {str(e)}")
                        # 오류 발생 시 해당 행의 옵션 필드를 빈 값으로 설정
                        processed_df.at[idx, 'option_mandatory'] = ''
                        processed_df.at[idx, 'option_mix'] = ''
                        processed_df.at[idx, 'option_name'] = ''
                        processed_df.at[idx, 'option_values'] = ''
                        processed_df.at[idx, 'option_prices'] = ''
                        processed_df.at[idx, 'opt_use'] = ''
                        processed_df.at[idx, 'opt_oneclick'] = ''
                        processed_df.at[idx, 'option_type'] = ''
            
            # 3. 나머지 필드 처리
            processed_df['selling_price'] = df.iloc[:, 11]
            processed_df['stock'] = df.iloc[:, 12]
            processed_df['consumer_price'] = df.iloc[:, 13]
            processed_df['supply_price'] = df.iloc[:, 14]
            processed_df['point_percentage'] = df.iloc[:, 15]
            processed_df['search_tags'] = df.iloc[:, 19]
            processed_df['adult_auth'] = 'Y'
            processed_df['display_status'] = '진열'
            processed_df['main_image'] = df.iloc[:, 22]
            processed_df['detail_image'] = 'AUTO'
            processed_df['list_image'] = 'AUTO'
            processed_df['mobile_image'] = df.iloc[:, 25]
            processed_df['description'] = df.iloc[:, 26]
            processed_df['mobile_description'] = df.iloc[:, 27]
            processed_df['model_name'] = df.iloc[:, 28]
            processed_df['best_product_display'] = 'N'
            processed_df['vat_type'] = 'Y'
            
            return processed_df
                
        except Exception as e:
            self.logger.error(f"데이터 전처리 중 오류: {str(e)}")
            self.logger.error(f"DataFrame 정보:\n{df.info()}")
            raise

    def convert(self) -> bool:
        """CSV를 Excel로 변환"""
        try:
            # CSV 읽기
            df = self.read_csv()
            if df.empty:
                raise Exception("CSV 데이터가 비어있습니다.")
            
            # 데이터 전처리
            processed_df = self.prepare_data(df)
            
            # 엑셀 템플릿 복사
            output_file = f'products_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            if os.path.exists(self.excel_template_path):
                # 엑셀 템플릿 로드
                workbook = openpyxl.load_workbook(self.excel_template_path)
                worksheet = workbook.active
                
                # 데이터 입력 시작 행 (헤더 2줄 다음부터)
                start_row = 3
                
                # 각 행 데이터 입력
                for idx, row in processed_df.iterrows():
                    current_row = start_row + idx
                    
                    # 필드 매핑에 따라 데이터 입력
                    for field, col_idx in self.field_mapping.items():
                        try:
                            value = row.get(field, '')
                            if pd.notna(value):  # null이 아닌 경우만 입력
                                cell = worksheet.cell(row=current_row, column=col_idx)
                                
                                # 숫자 데이터 처리
                                if isinstance(value, (int, float)):
                                    cell.value = value
                                else:
                                    cell.value = str(value)
                                    
                                # 셀 스타일 복사 (템플릿의 세 번째 행에서)
                                template_cell = worksheet.cell(row=3, column=col_idx)
                                if template_cell.has_style:
                                    cell._style = template_cell._style
                        except Exception as e:
                            self.logger.warning(f"셀 입력 중 오류 (행: {current_row}, 열: {col_idx}): {str(e)}")
                            continue
                
                # 파일 저장
                workbook.save(output_file)
                self.logger.info(f"Excel 파일 생성 완료: {output_file}")
                return True
                
            else:
                raise Exception("Excel 템플릿 파일을 찾을 수 없습니다.")
            
        except Exception as e:
            self.logger.error(f"Excel 변환 실패: {str(e)}")
            return False

def main():
    # 로깅 설정
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.StreamHandler(),  # 콘솔 출력
            logging.FileHandler('converter.log')  # 파일 출력
        ]
    )
    
    # 변환 실행
    converter = CSVtoExcelConverter()
    if converter.convert():
        print("CSV 파일이 성공적으로 Excel로 변환되었습니다.")
    else:
        print("변환 중 오류가 발생했습니다. 로그 파일을 확인해주세요.")

if __name__ == "__main__":
    main()